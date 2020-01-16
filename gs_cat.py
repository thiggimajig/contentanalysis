#!/usr/bin/env python3

"""
    gs_cat.py is used to create two charts and eventually a third 
    (1.Safe Urls 2.Most Popular Segments) 
    from the url data in the csv files created from the internal tool 
    called bulkcat (ie. grapeshot's url categorization tool). Each row 
    in the csv file represents a URL with it's corresponding grapeshot
    URL categorization data.
    Author: Taylor Higgins taylor.higgins@oracle.com
    Last modified: 1/12/2020

"""

#Todo: ask Shane for package manager for pandas etc. and his interface
#Todo: tell shane, recommendataions that didn't work.
#To do: create a requirements.txt file for openpyxl == 1.3.4 etc
#To do: change filename to name = sys.argv[1] so usable by all
#To do: save to a dynamic xls file name
#To do: if end up doing flask app and for 3rd more complicated have a sheet where they input match then save data on server then do calculations and display charts

import sys
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint

#ToDo: Move this to gate or make a function
#Open csv file from Bulkcat.
workbook = load_workbook(filename="WSJ_URL_Results_112519.xlsx")

#Add sheets for URL sections and charts. 
workbook.create_sheet('URL Sections')
workbook.create_sheet('Charts')
workbook.save('gs_cat_wsj.xlsx')

#Create new sheets for sections and charts and make variables of each sheet we'll need. 
cat_sheet = workbook['categorized']
sections_sheet = workbook['URL Sections']
charts_sheet = workbook['Charts']

#Add in headers and clean up the categorized sheet.
cat_sheet.delete_cols(2,3)
cat_sheet.insert_rows(1)
cat_sheet["A1"] = "URL"
cat_sheet["B1"] = "SEGMENT"
cat_sheet["C1"] = "SCORE"
cat_sheet["D1"] = "KEYWORDS"
cat_sheet["E1"] = "SEGMENT"
cat_sheet["F1"] = "SCORE"
cat_sheet["G1"] = "KEYWORDS"
cat_sheet["H1"] = "SEGMENT"
cat_sheet["I1"] = "SCORE"
cat_sheet["J1"] = "KEYWORDS"
cat_sheet["K1"] = "SEGMENT"
cat_sheet["L1"] = "SCORE"
cat_sheet["M1"] = "KEYWORDS"

#Add in headers for the sections sheet. 
sections_sheet['A1'] = "URL SECTION ONE"
sections_sheet['B1'] = "URL SECTION TWO"
sections_sheet['C1'] = "URL SECTION THREE"
sections_sheet['D1'] = "URL SECTION FOUR"
sections_sheet['E1'] = "URL SECTION FIVE"
sections_sheet['F1'] = "URL SECTION SIX"


def url_totals():

    """Find Totals for the charts. Categorized URLs total comes from the total row count in the categorized sheet 
    or by going through all the non gv_ and gx_ segments in the all sheet. Unsafe URLs Total
    comes from any url starting with gv_. """             

    unsafe_total = 0
    safe_segment_dict = {}

    #Starting loop to find uniquely unsafe URL rows.
    for row in cat_sheet.iter_rows(min_col=2, min_row=2, max_col=cat_sheet.max_column, max_row=cat_sheet.max_row): 
        for cell in row:
            unsafe_segment = re.compile(r'^gv_')
            try:
                #print(type(cell.value)) #str
                check_safety = unsafe_segment.search(cell.value)
                #ToDo change is is not None
                if check_safety != None: 
                    unsafe_total +=1 
                    #We put a break here so that we don't over count URLs that were categorized multiple times as unsafe.
                    break 
                    #We could add in an else to count for total segment appearnaces here.    
            except TypeError:
                #This should pass from the blank cell to the next in the row, then finally to next URL row.
                pass 
    #print(unsafe_total)

    #Starting loop to create a dict of absolute count of segment appearances regardless of unique URL row. 
    for row in cat_sheet.iter_rows(min_col=2, min_row=2, max_col=cat_sheet.max_column, max_row=cat_sheet.max_row):
        for cell in row:
            safe_segment = re.compile(r'^gs_')
            try:
                check_safety = safe_segment.search(cell.value)
                #Todo change again to is not
                #Todo do the default dict collections from Shane, ask him again
                #Todo: Figure out how to use a defaultdict here. 

                if check_safety != None:
                    new_segment = cell.value
                    if new_segment in safe_segment_dict: 
                        #If the segment is in the dict already add one to the current value.
                        safe_segment_dict[new_segment] = safe_segment_dict[new_segment] + 1
                    else:
                        #If the segment is not in the dict already then add it and make the value 1.
                        safe_segment_dict[new_segment] = 1
            except TypeError:
                pass
    #print(safe_segment_dict) 
    safe_segment_list = list(safe_segment_dict.items())

    #Calculate total categorized URLs. 
    row_count = cat_sheet.max_row 
    total_cat_urls = row_count - 1 #Make simpler with cat_sheet.max_row -1
    safe_total = total_cat_urls - unsafe_total

    charts_sheet['A1'] = 'Total URLs Categorized' 
    charts_sheet['B1'] = total_cat_urls

    return(unsafe_total, safe_total, safe_segment_list)


def safe_pie_chart(unsafe_total, safe_total):
    """Create Safe Unsafe Pie Chart"""

    #Here we create the data table and appending it in the Charts Sheet.
    #Todo: since only two items just call append twice
    safe_unsafe_data = [['Safe', safe_total], ['Unsafe', unsafe_total]]
    for row in safe_unsafe_data:
        charts_sheet.append(row)

    #Here we set up the pie chart using openpyxl. 
    pie = PieChart()
    labels = Reference(charts_sheet, min_col=1, min_row=2, max_row=3)
    data = Reference(charts_sheet, min_col=2, min_row=1, max_row=3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = 'Unsafe URls'
    charts_sheet.add_chart(pie, "H2")

def popular_bar_chart(safe_segment_list):
    """Create Popular Segments Bar Chart"""

    #To do: format in a little box around this to separate
    charts_sheet['A8'] = 'Segment'
    charts_sheet['B8'] = 'Total Appearances'

    #Here we append the total segment data to the Charts Sheet.
    for row in safe_segment_list:
        charts_sheet.append(row)

    #here we set up a bar chart using openpyxl.
    bar = BarChart()
    #fix lablels or pick less so can see names, something wrong with little squares
    #max row of 20
    labels = Reference(charts_sheet, min_col=1, min_row=4, max_row=111)
    data = Reference(worksheet=charts_sheet, min_row=4, max_row=111, min_col=2)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'Total Appearances'
    charts_sheet.add_chart(bar, "H20")


def url_path_parser():
    """Loop to parse all the URL path sections from the existing URLs. This will be used for the third chart. """
    #Todo: remove intermediary tuple section
    for row in cat_sheet.iter_rows(min_col=1, min_row=2, max_col=1, max_row=cat_sheet.max_row):    
        for cell in row:
            #print(type(cell.value)) #string
            split_cell_list = cell.value.split('/')[3:]
            tuple_section = (split_cell_list,)
            #print(type(tuple_section)) #tuple
            for row in tuple_section:
                #print(type(row)) #list
                sections_sheet.append(row)

def increased_content_percent():
    """This finds the % of x content that sits outside of x section. """
    #This data structure will hold the final tally of sections that would have a new url added if it went by segments.
    final_section_dict = {} #make it a defaultdict
    #This data structure loops through all rows and pulls out all unique sections.  
    all_sections_list = []
    #This data structure be created by looping through each row and for each url create a dictionary 
    #where the url is the key and the value is two lists. One list is a list of the sections for that url, 
    #the other list is a list of the segments for that url.
    url_info_dict = {url_one:[[url_one_sections],[url_one_segments]]}

def increased_content_percent_bar_chart():
    """This creates a bar chart of the top 20 website sections that would see the largest increase
    of url's after the url categorization tool. """


#Todo: change to gate or in init definition
#Todo: Out of curiousity figure out how you can add charts_ to the end of the name without disrupting xlsx. 
workbook.save(filename="gs_cat_wsj.xlsx")

if __name__ == '__main__':            
    url_path_parser()
    (unsafe_total, safe_total, safe_segment_list) = url_totals()
    safe_pie_chart(unsafe_total, safe_total)
    popular_bar_chart(safe_segment_list)



# """#!/usr/bin/env python3

# """
#     gs_cat.py is used to create two charts and eventually a third 
#     (1.Safe Urls 2.Most Popular Segments) 
#     from the url data in the csv files created from the internal tool 
#     called bulkcat (ie. grapeshot's url categorization tool). Each row 
#     in the csv file represents a URL with it's corresponding grapeshot
#     URL categorization data.
#     Author: Taylor Higgins taylor.higgins@oracle.com
#     Last modified: 1/15/2020

# """
# import sys
# import re
# import openpyxl
# import operator
# import collections
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
# from openpyxl.chart import BarChart, Reference, PieChart
# from openpyxl.chart.series import DataPoint

# def url_path_parser():
#     """Loop to parse all the URL path sections from the existing URLs. This will be used for the third chart which I haven't started yet. """
#     for row in cat_sheet.iter_rows(min_col=1, min_row=2, max_col=1, max_row=cat_sheet.max_row):    
#         for cell in row:
            
#             split_cell_list = cell.value.split('/')[3:]
#             tuple_section = (split_cell_list,)
            
#             for row in tuple_section:
                
#                 sections_sheet.append(row)

# def url_totals():

#     """Find Totals for the charts. Categorized URLs total comes from the total row count in the categorized sheet 
#     or by going through all the non gv_ and gx_ segments in the all sheet. Unsafe URLs Total
#     comes from any url starting with gv_. """             

#     unsafe_total = 0
#     safe_segment_dict = {}

#     #Starting loop to find uniquely unsafe URL rows.
#     for row in cat_sheet.iter_rows(min_col=2, min_row=2, max_col=cat_sheet.max_column, max_row=cat_sheet.max_row): 
#         for cell in row:
#             unsafe_segment = re.compile(r'^gv_')
#             try:
                
#                 check_safety = unsafe_segment.search(cell.value)
#                 if check_safety is not None: 
#                     unsafe_total +=1 
#                     #We put a break here so that we don't over count URLs that were categorized multiple times as unsafe.
#                     break 

#             except TypeError:
#                 #This should pass from the blank cell to the next in the row, then finally to next URL row.
#                 pass 

#     #Starting loop to create a dict of absolute count of segment appearances regardless of unique URL row. 
#     for row in cat_sheet.iter_rows(min_col=2, min_row=2, max_col=cat_sheet.max_column, max_row=cat_sheet.max_row):
#         for cell in row:
#             safe_segment = re.compile(r'^gs_')
#             try:
#                 check_safety = safe_segment.search(cell.value)
#                 if check_safety is not None:
#                     new_segment = cell.value
#                     if new_segment in safe_segment_dict: 
#                         #If the segment is in the dict already add one to the current value.
#                         safe_segment_dict[new_segment] = safe_segment_dict[new_segment] + 1
#                     else:
#                         #If the segment is not in the dict already then add it and make the value 1.
#                         safe_segment_dict[new_segment] = 1
#             except TypeError:
#                 pass
#     #Here we are sorting the dict of safe segments in descending order by the value of the dict.   
#     sorted_safe_seg_dict = dict(sorted(safe_segment_dict.items(), key=operator.itemgetter(1), reverse=True))
    
#     #Here we are turning the sorted dict into a list so we can later append it by row and create the segment bar chart.
#     safe_segment_list = list(sorted_safe_seg_dict.items())

#     #Calculate total categorized URLs. 
#     row_count = cat_sheet.max_row 
#     total_cat_urls = row_count - 1 
#     safe_total = total_cat_urls - unsafe_total

#     charts_sheet['A1'] = 'Total URLs Categorized' 
#     charts_sheet['B1'] = total_cat_urls

#     return(unsafe_total, safe_total, safe_segment_list)

# def safe_pie_chart(unsafe_total, safe_total):
#     """Create Safe Unsafe Pie Chart"""

#     #Here we create the data table and appending it in the Charts Sheet.
#     safe_unsafe_data = [['Safe', safe_total], ['Unsafe', unsafe_total]]
#     for row in safe_unsafe_data:
#         charts_sheet.append(row)

#     #Here we set up the pie chart using openpyxl. 
#     pie = PieChart()
#     labels = Reference(charts_sheet, min_col=1, min_row=2, max_row=3)
#     data = Reference(charts_sheet, min_col=2, min_row=1, max_row=3)
#     pie.add_data(data, titles_from_data=True)
#     pie.set_categories(labels)
#     pie.title = 'Unsafe URls'
#     charts_sheet.add_chart(pie, "H2")

# def popular_bar_chart(safe_segment_list):
#     """Create Popular Segments Bar Chart"""

#     charts_sheet['A6'] = 'Segment'
#     charts_sheet['B6'] = 'Total Appearances'

#     #Here we append the total segment data to the Charts Sheet.
#     for row in safe_segment_list:
#         charts_sheet.append(row)

#     #Here we set up a bar chart using openpyxl.
#     bar = BarChart()
#     labels = Reference(charts_sheet, min_col=1, min_row=7, max_row=26)
#     data = Reference(worksheet=charts_sheet, min_col=2, min_row=7, max_row=26)
#     bar.add_data(data, titles_from_data=True)
#     bar.set_categories(labels)
#     bar.title = 'Total Appearances'
#     charts_sheet.add_chart(bar, "H20")

# if __name__ == '__main__': 
#     #Open csv file from Bulkcat using the command line argument.
#     name = sys.argv[1]
#     workbook = load_workbook(filename=name) #WSJ_URL_Results_112519.xlsx
#     #Add sheets for URL sections and charts. 
#     workbook.create_sheet('URL Sections')
#     workbook.create_sheet('Charts')
#     #Create new sheets for sections and charts and make variables of each sheet we'll need. 
#     cat_sheet = workbook['categorized']
#     sections_sheet = workbook['URL Sections']
#     charts_sheet = workbook['Charts']
#     #Add in headers and clean up the categorized sheet.
#     cat_sheet.delete_cols(2,3)
#     cat_sheet.insert_rows(1)
#     cat_sheet["A1"] = "URL"
#     cat_sheet["B1"] = "SEGMENT"
#     cat_sheet["C1"] = "SCORE"
#     cat_sheet["D1"] = "KEYWORDS"
#     cat_sheet["E1"] = "SEGMENT"
#     cat_sheet["F1"] = "SCORE"
#     cat_sheet["G1"] = "KEYWORDS"
#     cat_sheet["H1"] = "SEGMENT"
#     cat_sheet["I1"] = "SCORE"
#     cat_sheet["J1"] = "KEYWORDS"
#     #Add in headers for the sections sheet. 
#     sections_sheet['A1'] = "URL SECTION ONE"
#     sections_sheet['B1'] = "URL SECTION TWO"
#     sections_sheet['C1'] = "URL SECTION THREE"
#     url_path_parser()
#     (unsafe_total, safe_total, safe_segment_list) = url_totals()
#     safe_pie_chart(unsafe_total, safe_total)
#     popular_bar_chart(safe_segment_list)
#     workbook.save(filename="{}{}".format("charts_",name))
# """
