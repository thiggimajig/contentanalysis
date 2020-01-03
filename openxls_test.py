#!/usr/bin/env python3
import sys
import re
import numpy as np
import pandas as pd
import seaborn as sns
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color

#creates workbook object from file
workbook = load_workbook(filename="wsj_test.xlsx")

#addign my section tab/sheet
workbook.create_sheet('URL Sections')
workbook.save('wsj_pretty_section.xlsx')

#this determines which tab to work on
cat_sheet = workbook['categorized']
sections_sheet = workbook['URL Sections']

#need v2.5+ to add and delete rows cols
cat_sheet.delete_cols(2,3)
cat_sheet.insert_rows(1)
cat_sheet["A1"] = "URL"
cat_sheet["B1"] = "SEGMENT"
cat_sheet["C1"] = "SCORE"
cat_sheet["D1"] = "KEYWORDS"
cat_sheet["E1"] = "SEGMENT"
cat_sheet["F1"] = "SCORE"
cat_sheet["G1"] = "KEYWORDS"
cat_sheet["H1"] = "SEGMENT"
cat_sheet["I1"] = "SCORE"
cat_sheet["J1"] = "KEYWORDS"
cat_sheet["K1"] = "SEGMENT"
cat_sheet["L1"] = "SCORE"
cat_sheet["M1"] = "KEYWORDS"

sections_sheet['A1'] = "URL SECTION ONE"
sections_sheet['B1'] = "URL SECTION TWO"
sections_sheet['C1'] = "URL SECTION THREE"
sections_sheet['D1'] = "URL SECTION FOUR"


#break out url by subdomain "/" and put them in a column
#iterate through B column and pull out text between "/"
#regex: \/(.*?)\/
#easier way is to use split on / but then skip first 3
#adding it to own tab because of iter and makes sense

#once we have the sections for each url and segment
#we can do the work for the second chart
#do that by seeing if segments for a url match to any known sections
#if the segment word does for a given url and it's not in that section already
#then we add a tally for that section as missed opp
#each tally for a given section goes to that bar chart
#find the percent that urls for a given section do not sit in that section 
#add up all url's for a section, in and out, based on segment (and section?)
#show percent of those url's that should be in section based on segment word but aren't  

for row in cat_sheet.iter_rows(min_col=2, min_row=2, max_col=2, max_row=cat_sheet.max_row):    
    for cell in row:
        #print(cell.value)
        #print(type(cell.value)) #string
        #split and only retrieve items after protocol and domain ie. the path sections
        split_cell_list = cell.value.split('/')[3:]
        #let's try tuples
        tuple_section = (split_cell_list,)
        #print(tuple_section)
        #print(type(tuple_section)) #tuple
        for row in tuple_section:
            #print(row)
            #print(type(row)) #list
            sections_sheet.append(row)


#1. count total categorized urls in ALL column of ugly by non gx or by counting all urls in categorized tab
#2. count total unsafe by going through segment in categorized and if it has gv
#3. count total safe by going through segment in categorized and if it has gs
#4. count total segment appearances in general in cat
#5. count total of segment appearances per unique url in cat

#easy charts
#1. safe/unsafe pie chart show % in a pie chart
#2. total appearances bar chart, show highest 15 in B7

#hard charts
#1. count url number (% of total) of a segment that isn't in that segment's subdomain section 



#FORMATTING STUFF LATER UNIMPORTANT
#unsure create font and fill to apply to header
# font = Font(name='Calibri', size=77, bold=True, color='ffffff')
# sheet["A1"].font = font
#fill = PatternFill(fill_type='solid', color='ff0000') 
#unsure loop through all columns in first row to apply font
#for icell in row 1 icell.font=font 
#saves our work to new file

#SUMMARY TAB UNIMPORTANT FOR NOW
# I don't think they care about this only the charts
# #workbook.sheetnames[1]
# sheet_summary = workbook.active
# #first bit
# sheet["A2"] = "Successfully Categorized (total)"
# sheet["A3"] = "Categorized w/ CUSTOM Taxonomy"
# sheet["A4"] = "Unsafe URLs"
# sheet["B1"] = "Total" #font
# sheet["C1"] = "%" #font
# #second bit
# sheet["A7"] = "Segment" #font
# sheet["B7"] = "Total Appearances" #font
# sheet["E7"] = "Segment" #font
# sheet["F7"] = "# URLs" #font
# sheet["G7"] = "% URLs" #font
# sheet["J7"] = "Custom Segments" #font
# sheet["K7"] = "# URLs" #font
# sheet["L7"] = "% URLs" #font

workbook.save(filename="wsj_pretty_section.xlsx")


